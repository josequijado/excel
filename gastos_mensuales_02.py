import random
import xlsxwriter
from openpyxl import load_workbook


class GastosMensuales:
    def __init__(self, umbral=150.0):
        self.meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        self.partidas = ["Luz", "Agua", "Internet", "Comunidad"]
        self.umbral = umbral
        self.datos = self._generar_datos()

    def _generar_datos(self):
        return [[round(random.uniform(50, 200), 2) for _ in self.partidas] for _ in self.meses]

    def _calcular_totales(self):
        totales_por_columna = [round(sum(col), 2) for col in zip(*self.datos)]
        total_general = round(sum(totales_por_columna), 2)
        return totales_por_columna, total_general

    def _ajustar_columnas(self, worksheet):
        # Ajustar manualmente el ancho de la columna "Mes"
        worksheet.set_column(0, 0, 20)  # Ampliar el ancho de la columna de los meses

        # Ajustar automáticamente el resto de las columnas con margen adicional
        for col_idx, column in enumerate(["Mes"] + self.partidas, start=0):
            worksheet.set_column(col_idx, col_idx, len(column) + 8)  # Ancho ajustado con margen

    def _agregar_filtros(self, nombre_archivo):
        # Abrir el archivo con openpyxl para agregar filtros
        wb = load_workbook(nombre_archivo)
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save(nombre_archivo)
        print("Filtros habilitados en el archivo.")

    def generar_excel(self, nombre_archivo="gastos_mensuales_02.xlsx"):
        workbook = xlsxwriter.Workbook(nombre_archivo)
        worksheet = workbook.add_worksheet("Gastos Mensuales")

        # Formatos
        header_format = workbook.add_format({"bold": True, "bg_color": "#00FF99", "font_size": 16})
        data_format = workbook.add_format({"font_size": 14})
        red_bold_format = workbook.add_format({"bold": True, "font_color": "red", "font_size": 14})

        # Escribir encabezados
        worksheet.write_row(0, 0, ["Mes"] + self.partidas, header_format)

        # Escribir datos
        for row_idx, (mes, valores) in enumerate(zip(self.meses, self.datos), start=1):
            worksheet.write(row_idx, 0, mes, data_format)
            for col_idx, valor in enumerate(valores, start=1):
                if valor > self.umbral:
                    worksheet.write(row_idx, col_idx, valor, red_bold_format)
                else:
                    worksheet.write(row_idx, col_idx, valor, data_format)

        # Escribir totales
        totales, total_general = self._calcular_totales()
        worksheet.write(len(self.meses) + 1, 0, "Total", header_format)
        for col_idx, total in enumerate(totales, start=1):
            worksheet.write(len(self.meses) + 1, col_idx, total, data_format)
        worksheet.write(len(self.meses) + 1, len(totales) + 1, total_general, data_format)

        # Ajustar el ancho de las columnas
        self._ajustar_columnas(worksheet)

        workbook.close()
        print(f"Archivo Excel generado: {nombre_archivo}")

        # Agregar filtros con openpyxl
        self._agregar_filtros(nombre_archivo)


if __name__ == "__main__":
    gastos = GastosMensuales(umbral=150.0)
    gastos.generar_excel()
